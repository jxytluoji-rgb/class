import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import io

# 节次时间（按照最开始做的程序）
SECTION_TIMES = {
    1: ("第一节", "08:00", "08:40"),
    2: ("第二节", "08:50", "09:30"),
    3: ("第三节", "09:45", "10:25"),
    4: ("第四节", "10:40", "11:20"),
    5: ("第五节", "11:30", "12:10"),
    6: ("第六节", "14:00", "14:40"),
    7: ("第七节", "14:50", "15:30"),
    8: ("第八节", "15:40", "16:20"),
    9: ("第九节", "16:30", "17:10"),
}

# 教室列表（701-706, 708-713, 801-806, 808-813）
CLASSROOMS = []
for floor in ['7', '8']:
    for room in range(1, 14):
        classroom = f"{floor}{room:02d}"
        if classroom not in ['707', '807']:
            CLASSROOMS.append(classroom)

# 设备ID映射
DEVICE_IDS = {
    '701': 'dc-62-94-0d-d9-88',
    '702': 'dc-62-94-0d-d9-d7',
    '703': 'dc-62-94-0a-8d-db',
    '704': 'dc-62-94-0d-da-93',
    '705': 'dc-62-94-0d-d9-7c',
    '706': 'dc-62-94-0d-d8-44',
    '708': 'dc-62-94-0d-d8-a8',
    '709': 'dc-62-94-0d-d7-32',
    '710': 'dc-62-94-0d-d8-55',
    '711': 'dc-62-94-0d-dc-56',
    '712': 'dc-62-94-0d-d9-13',
    '713': 'dc-62-94-0d-d9-d8',
    '801': 'dc-62-94-0d-d9-96',
    '802': 'dc-62-94-0d-d9-cf',
    '803': 'dc-62-94-0d-d8-2c',
    '804': 'dc-62-94-0d-d7-30',
    '805': 'dc-62-94-0d-da-aa',
    '806': 'dc-62-94-0d-d9-81',
    '808': 'dc-62-94-0d-d9-a2',
    '809': 'dc-62-94-0d-d9-7d',
    '810': 'dc-62-94-0d-db-47',
    '811': 'dc-62-94-0d-da-83',
    '812': 'dc-62-94-0d-d9-91',
    '813': 'dc-62-94-0d-d9-7f'
}

# 星期列表
WEEKDAYS = ['一', '二', '三', '四', '五']

def parse_course_info(course_str):
    if pd.isna(course_str) or course_str == '' or course_str == 'nan':
        return ""
    course_str = str(course_str).strip()
    course_str = re.sub(r'\[.*?\]', '', course_str).strip()
    return course_str

def parse_teacher_info(teacher_str):
    if pd.isna(teacher_str) or teacher_str == '' or teacher_str == 'nan':
        return "", ""
    teacher_str = str(teacher_str).strip()
    match = re.search(r'\[(\d+)\](.+)', teacher_str)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    parts = teacher_str.split('-')
    if len(parts) >= 2:
        return parts[0].strip(), '-'.join(parts[1:]).strip()
    return "", teacher_str

def parse_class_info(class_str):
    if pd.isna(class_str) or class_str == '' or class_str == 'nan':
        return ""
    class_str = str(class_str).strip()
    if "产业园C栋" in class_str or "容量" in class_str:
        return ""
    class_matches = re.findall(r'(\d+产)\s*[（](\d+)[）]', class_str)
    if class_matches:
        class_name = " ".join([f"{m[0]}{m[1]}" for m in class_matches])
    else:
        class_name = class_str
    return class_name

def parse_classroom_info(class_str):
    if pd.isna(class_str) or class_str == '' or class_str == 'nan':
        return None
    class_str = str(class_str).strip()
    if "产业园C栋" in class_str and "容量" in class_str:
        match = re.search(r'产业园C栋(\d+)', class_str)
        if match:
            room_num = match.group(1)
            if room_num not in ['508', '509', '901']:
                if room_num.startswith('7'):
                    building = "产业园7楼"
                elif room_num.startswith('8'):
                    building = "产业园8楼"
                else:
                    building = "产业园"
                return (f"产业园C栋{room_num}", building, "产业园")
    return None

def parse_week_info(week_str):
    if pd.isna(week_str) or week_str == '' or week_str == 'nan':
        return []
    week_str = str(week_str).strip()
    weeks = []
    if '-' in week_str:
        parts = week_str.split('-')
        if len(parts) == 2:
            try:
                start = int(parts[0])
                end = int(parts[1])
                weeks = list(range(start, end + 1))
            except ValueError:
                pass
    else:
        try:
            weeks = [int(week_str)]
        except ValueError:
            pass
    return weeks

def parse_section_info(section_str):
    if pd.isna(section_str) or section_str == '' or section_str == 'nan':
        return {"星期": "", "节次": ""}
    section_str = str(section_str).strip()
    match = re.match(r'([一二三四五六日])\[(.+?)\](.*)', section_str)
    if match:
        section_content = match.group(2)
        suffix = match.group(3).strip()
        if suffix:
            section_content += suffix
        return {
            "星期": match.group(1),
            "节次": section_content
        }
    return {"星期": "", "节次": section_str}

def parse_section_range(section_str):
    if pd.isna(section_str) or section_str == '' or section_str == 'nan':
        return [], None
    section_str = str(section_str).strip()
    week_type = None
    if '单' in section_str:
        week_type = '单'
        section_str = section_str.replace('单', '')
    elif '双' in section_str:
        week_type = '双'
        section_str = section_str.replace('双', '')
    section_str = section_str.replace('节', '')
    sections = []
    if '-' in section_str:
        parts = section_str.split('-')
        if len(parts) == 2:
            try:
                start = int(parts[0])
                end = int(parts[1])
                sections = list(range(start, end + 1))
            except ValueError:
                pass
    else:
        try:
            sections = [int(section_str)]
        except ValueError:
            pass
    return sections, week_type

def parse_teacher_info_classroom(teacher_str):
    if pd.isna(teacher_str) or teacher_str == '' or teacher_str == 'nan':
        return ""
    teacher_str = str(teacher_str).strip()
    match = re.search(r'\[(\d+)\](.+)', teacher_str)
    if match:
        return match.group(2).strip()
    parts = teacher_str.split('-')
    if len(parts) >= 2:
        return '-'.join(parts[1:]).strip()
    return teacher_str

def parse_section_info_classroom(section_str):
    if pd.isna(section_str) or section_str == '' or section_str == 'nan':
        return {"星期": "", "节次": "", "单双周": ""}
    section_str = str(section_str).strip()
    match = re.match(r'([一二三四五六日])\[(.+?)\](.*)', section_str)
    if match:
        section_content = match.group(2)
        suffix = match.group(3).strip()
        return {
            "星期": match.group(1),
            "节次": section_content,
            "单双周": suffix
        }
    return {"星期": "", "节次": "", "单双周": ""}

def parse_section_range_classroom(section_str):
    sections = []
    if pd.isna(section_str) or section_str == '' or section_str == 'nan':
        return sections

    section_str = str(section_str).strip()
    section_str = section_str.replace('节', '')

    parts = section_str.split(',')
    for part in parts:
        part = part.strip()
        if '-' in part:
            try:
                start, end = part.split('-')
                start = int(start)
                end = int(end)
                sections.extend(range(start, end + 1))
            except ValueError:
                pass
        else:
            try:
                sections.append(int(part))
            except ValueError:
                pass

    return sections

def convert_classroom_schedule(input_file, output_file, semester_start_date="2026-03-09", target_week=None):
    df_source = pd.read_excel(input_file, header=4)
    df_source = df_source.iloc[1:]
    target_data = []
    semester_start = datetime.strptime(semester_start_date, "%Y-%m-%d")
    course_serial_map = {}
    current_serial = 1
    daily_course_count = {}
    current_classroom = None
    current_building = None
    current_campus = None

    for idx, row in df_source.iterrows():
        classroom_info = parse_classroom_info(str(row.get('上课班级构成', '')))
        if classroom_info:
            current_classroom, current_building, current_campus = classroom_info
            continue
        elif "产业园C栋" in str(row.get('上课班级构成', '')) and "容量" in str(row.get('上课班级构成', '')):
            current_classroom = None
            current_building = None
            current_campus = None
            continue
        if not current_classroom:
            continue

        course_name = parse_course_info(str(row.get('课程', '')))
        teacher_code, teacher_name = parse_teacher_info(str(row.get('教师', '')))
        class_name = parse_class_info(str(row.get('上课班级构成', '')))
        week_value = str(row.get('周次', ''))
        if week_value == '周次' or week_value == '节次' or week_value == 'nan':
            continue

        weeks = parse_week_info(week_value)
        section_info = parse_section_info(str(row.get('节次', '')))
        weekday = section_info["星期"]
        section = section_info["节次"]

        if not weeks or not weekday or not section:
            continue

        section_numbers, week_type = parse_section_range(section)
        if not section_numbers:
            continue

        if course_name not in course_serial_map:
            course_serial_map[course_name] = current_serial
            current_serial += 1
        course_serial = course_serial_map[course_name]

        for week in weeks:
            if target_week is not None and week != target_week:
                continue
            if week_type == '单' and week % 2 == 0:
                continue
            if week_type == '双' and week % 2 == 1:
                continue

            weekday_map = {"一": 0, "二": 1, "三": 2, "四": 3, "五": 4, "六": 5, "日": 6}
            weekday_num = weekday_map.get(weekday, 0)
            course_date = semester_start + timedelta(weeks=week-1, days=weekday_num)
            date_str = course_date.strftime("%Y-%m-%d")
            date_short = course_date.strftime("%Y-%m-%d")

            for section_num in section_numbers:
                if section_num in SECTION_TIMES:
                    section_name, start_time, end_time = SECTION_TIMES[section_num]
                else:
                    continue

                daily_key = f"{date_str}-{course_name}"
                if daily_key not in daily_course_count:
                    daily_course_count[daily_key] = 1
                else:
                    daily_course_count[daily_key] += 1
                class_count = daily_course_count[daily_key]
                record_id = f"{course_name}-{date_short}-{class_count}"
                classroom_number = current_classroom.replace("产业园C栋", "")

                target_record = {
                    "编号": record_id,
                    "教师账号": teacher_code,
                    "课程及编号": f"{course_name}，{course_serial}",
                    "校区": current_campus,
                    "教学楼": current_building,
                    "教室": classroom_number,
                    "节次类型": "产业学院冬季",
                    "节次名称": section_name,
                    "开始时间": f"{date_str} {start_time}" if start_time else "",
                    "结束时间": f"{date_str} {end_time}" if end_time else "",
                    "组织机构": "数字技术应用产业学院",
                    "班级": class_name,
                    "简介": "",
                    "是否需要录制": "TRUE",
                    "录制视频是否公开": "false",
                    "是否直播": "",
                    "直播是否公开": ""
                }
                target_data.append(target_record)

    df_target = pd.DataFrame(target_data)
    if not df_target.empty:
        df_target['教室排序'] = df_target['教室'].astype(int)
        df_target = df_target.sort_values(by=['教室排序', '开始时间']).reset_index(drop=True)
        df_target = df_target.drop(columns=['教室排序'])

    df_target.to_excel(output_file, index=False, engine='openpyxl')
    wb = load_workbook(output_file)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(output_file)
    return len(target_data)

def read_source_file_classroom(file_path, target_week):
    df = pd.read_excel(file_path, header=None)
    course_data = {}
    current_classroom = None

    for index, row in df.iterrows():
        if '产业园C栋' in str(row.iloc[3]) and '容量' in str(row.iloc[13]):
            classroom_str = str(row.iloc[13])
            match = re.search(r'产业园C栋(\d+)', classroom_str)
            if match:
                classroom_num = match.group(1)
                if (classroom_num.startswith('7') or classroom_num.startswith('8')) and len(classroom_num) == 3:
                    current_classroom = classroom_num
                else:
                    current_classroom = None
            else:
                current_classroom = None
            continue

        if index < 5:
            continue

        if not current_classroom:
            continue

        course_cell = row.iloc[1]
        if pd.isna(course_cell):
            continue

        course_str = str(course_cell).strip()
        teacher_id_match = re.search(r'\[(\d+)\]', course_str)
        if teacher_id_match:
            teacher_id = teacher_id_match.group(1)
            course_name = course_str.replace(f"[{teacher_id}]", "").strip()
        else:
            teacher_id = ""
            course_name = course_str

        course_name = re.sub(r'\[.*?\]', '', course_name).strip()

        teacher_cell = row.iloc[10]
        teacher_name = parse_teacher_info_classroom(teacher_cell)

        class_cell = row.iloc[13]
        class_name = parse_class_info(class_cell)

        week_str = row.iloc[15]
        week_info = parse_week_info(week_str)

        if target_week not in week_info:
            continue

        section_str = row.iloc[16]
        section_info = parse_section_info_classroom(section_str)

        week_type = section_info.get('单双周', '')
        if week_type:
            if week_type == '单' and target_week % 2 == 0:
                continue
            if week_type == '双' and target_week % 2 != 0:
                continue

        sections = parse_section_range_classroom(section_info.get('节次', ''))
        weekday = section_info.get('星期', '')

        for section in sections:
            if section < 1 or section > 9:
                continue

            key = (current_classroom, weekday, section)
            course_data[key] = {
                '班级名称': class_name,
                '设备id': '',
                '星期': weekday,
                '节次': section,
                '时段': f"{SECTION_TIMES.get(section, ('', '', ''))[1]}-{SECTION_TIMES.get(section, ('', '', ''))[2]}",
                '教师': teacher_name,
                '科目': course_name
            }

    return course_data

def generate_classroom_sign(course_data, output_path):
    result_data = []

    for classroom in CLASSROOMS:
        for weekday in WEEKDAYS:
            for section in range(1, 10):
                key = (classroom, weekday, section)
                if key in course_data:
                    record = course_data[key].copy()
                    record['班级名称'] = classroom
                    record['设备id'] = DEVICE_IDS.get(classroom, '')
                    if not record['教师']:
                        record['教师'] = '空闲'
                    if not record['科目']:
                        record['科目'] = '空闲'
                    result_data.append(record)
                else:
                    section_name, start_time, end_time = SECTION_TIMES.get(section, ('', '', ''))
                    result_data.append({
                        '班级名称': classroom,
                        '设备id': DEVICE_IDS.get(classroom, ''),
                        '星期': weekday,
                        '节次': section,
                        '时段': f"{start_time}-{end_time}",
                        '教师': '空闲',
                        '科目': '空闲'
                    })

    df = pd.DataFrame(result_data)

    df.columns = [
        '班级名称',
        '设备id',
        '星期(一、二、三、四、五、六、日,请按顺序排列)',
        '节次(1、2、3、4、5、6、7、8,请按顺序排列,如果当天中间无课程,请插入一条科目、教师为空白的记录)',
        '时段(09:00-09:45)',
        '教师',
        '科目'
    ]

    df.to_excel(output_path, index=False, sheet_name='Timetable')

    wb = load_workbook(output_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_path)

    return len(result_data)

def get_excel_download_link(file_path, filename, link_text):
    with open(file_path, 'rb') as f:
        file_data = f.read()
    b64 = base64.b64encode(file_data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{filename}">{link_text}</a>'
    return href
