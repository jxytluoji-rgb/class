from flask import Flask, request, render_template, send_file, jsonify
import os
import tempfile
import pandas as pd
from datetime import datetime, timedelta
import re
import json
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

# 创建上传目录
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# ==================== 配置文件管理 ====================
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')

# 默认配置
DEFAULT_CONFIG = {
    'semester_start_date': '2026-03-09',
    'convert_type': 'classroom_schedule',
    'season': 'winter',
    'week': 1
}

def load_config():
    """加载配置文件"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # 合并默认配置，确保所有键都存在
                return {**DEFAULT_CONFIG, **config}
        except:
            return DEFAULT_CONFIG.copy()
    return DEFAULT_CONFIG.copy()

def save_config(config):
    """保存配置文件"""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False

# ==================== 时间配置 ====================
# 冬季课表时间（10月2日-4月30日）
WINTER_SECTION_TIMES = {
    1: ("第一节", "08:00", "08:40"),
    2: ("第二节", "08:50", "09:30"),
    3: ("第三节", "09:45", "10:25"),
    4: ("第四节", "10:40", "11:20"),
    5: ("第五节", "11:30", "12:10"),
    6: ("第六节", "14:00", "14:40"),
    7: ("第七节", "14:50", "15:30"),
    8: ("第八节", "15:40", "16:20"),
    9: ("第九节", "16:30", "17:10"),
}

# 夏季课表时间（5月1日-10月1日）- 第6-9节课往后推延30分钟
SUMMER_SECTION_TIMES = {
    1: ("第一节", "08:00", "08:40"),
    2: ("第二节", "08:50", "09:30"),
    3: ("第三节", "09:45", "10:25"),
    4: ("第四节", "10:40", "11:20"),
    5: ("第五节", "11:30", "12:10"),
    6: ("第六节", "14:30", "15:10"),
    7: ("第七节", "15:20", "16:00"),
    8: ("第八节", "16:10", "16:50"),
    9: ("第九节", "17:00", "17:40"),
}

def get_section_details(season='winter'):
    if season == 'summer':
        return SUMMER_SECTION_TIMES
    return WINTER_SECTION_TIMES

# ==================== 教室课表文件功能 ====================

def parse_course_info(course_str):
    if pd.isna(course_str) or course_str == '' or course_str == 'nan':
        return ""
    course_str = str(course_str).strip()
    # 移除所有方括号及其内容
    course_str = re.sub(r'\[.*?\]', '', course_str).strip()
    return course_str

def parse_teacher_info(teacher_str):
    if pd.isna(teacher_str) or teacher_str == '' or teacher_str == 'nan':
        return "", ""
    teacher_str = str(teacher_str).strip()
    # 解析方括号中的教师账号
    match = re.search(r'\[(\d+)\](.+)', teacher_str)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    # 保持原有的解析逻辑作为后备
    parts = teacher_str.split('-')
    if len(parts) >= 2:
        return parts[0].strip(), '-'.join(parts[1:]).strip()
    return "", teacher_str

def parse_class_info(class_str):
    if pd.isna(class_str) or class_str == '' or class_str == 'nan':
        return ""
    class_str = str(class_str).strip()
    if "产业园C栋" in class_str or "容量" in class_str:
        return ""
    class_matches = re.findall(r'(\d+产)\s*[（](\d+)[）]', class_str)
    if class_matches:
        class_name = " ".join([f"{m[0]}{m[1]}" for m in class_matches])
    else:
        class_name = class_str
    return class_name

def parse_classroom_info(class_str):
    if pd.isna(class_str) or class_str == '' or class_str == 'nan':
        return None
    class_str = str(class_str).strip()
    if "产业园C栋" in class_str and "容量" in class_str:
        match = re.search(r'产业园C栋(\d+)', class_str)
        if match:
            room_num = match.group(1)
            if room_num not in ['508', '509', '901']:
                if room_num.startswith('7'):
                    building = "产业园7楼"
                elif room_num.startswith('8'):
                    building = "产业园8楼"
                else:
                    building = "产业园"
                return (f"产业园C栋{room_num}", building, "产业园")
    return None

def parse_week_info(week_str):
    if pd.isna(week_str) or week_str == '' or week_str == 'nan':
        return []
    week_str = str(week_str).strip()
    weeks = []
    if '-' in week_str:
        parts = week_str.split('-')
        if len(parts) == 2:
            try:
                start = int(parts[0])
                end = int(parts[1])
                weeks = list(range(start, end + 1))
            except ValueError:
                pass
    else:
        try:
            weeks = [int(week_str)]
        except ValueError:
            pass
    return weeks

def parse_section_info(section_str):
    if pd.isna(section_str) or section_str == '' or section_str == 'nan':
        return {"星期": "", "节次": ""}
    section_str = str(section_str).strip()
    match = re.match(r'([一二三四五六日])\[(.+?)\](.*)', section_str)
    if match:
        section_content = match.group(2)
        suffix = match.group(3).strip()
        if suffix:
            section_content += suffix
        return {
            "星期": match.group(1),
            "节次": section_content
        }
    return {"星期": "", "节次": section_str}

def parse_section_range(section_str):
    if pd.isna(section_str) or section_str == '' or section_str == 'nan':
        return [], None
    section_str = str(section_str).strip()
    week_type = None
    if '单' in section_str:
        week_type = '单'
        section_str = section_str.replace('单', '')
    elif '双' in section_str:
        week_type = '双'
        section_str = section_str.replace('双', '')
    section_str = section_str.replace('节', '')
    sections = []
    if '-' in section_str:
        parts = section_str.split('-')
        if len(parts) == 2:
            try:
                start = int(parts[0])
                end = int(parts[1])
                sections = list(range(start, end + 1))
            except ValueError:
                pass
    else:
        try:
            sections = [int(section_str)]
        except ValueError:
            pass
    return sections, week_type

def convert_classroom_schedule(input_file, output_file, semester_start_date="2026-03-09", target_week=None, week_type_filter=None, season='winter'):
    df_source = pd.read_excel(input_file, header=4)
    df_source = df_source.iloc[1:]
    target_data = []
    semester_start = datetime.strptime(semester_start_date, "%Y-%m-%d")
    course_serial_map = {}
    current_serial = 1
    daily_course_count = {}
    section_details = get_section_details(season)
    current_classroom = None
    current_building = None
    current_campus = None
    season_type = "产业学院夏季" if season == 'summer' else "产业学院冬季"
    
    for idx, row in df_source.iterrows():
        classroom_info = parse_classroom_info(str(row.get('上课班级构成', '')))
        if classroom_info:
            current_classroom, current_building, current_campus = classroom_info
            continue
        elif "产业园C栋" in str(row.get('上课班级构成', '')) and "容量" in str(row.get('上课班级构成', '')):
            current_classroom = None
            current_building = None
            current_campus = None
            continue
        if not current_classroom:
            continue
        course_name = parse_course_info(str(row.get('课程', '')))
        teacher_code, teacher_name = parse_teacher_info(str(row.get('教师', '')))
        class_name = parse_class_info(str(row.get('上课班级构成', '')))
        week_value = str(row.get('周次', ''))
        if week_value == '周次' or week_value == '节次' or week_value == 'nan':
            continue
        weeks = parse_week_info(week_value)
        section_info = parse_section_info(str(row.get('节次', '')))
        weekday = section_info["星期"]
        section = section_info["节次"]
        if not weeks or not weekday or not section:
            continue
        section_numbers, week_type = parse_section_range(section)
        if not section_numbers:
            continue
        if course_name not in course_serial_map:
            course_serial_map[course_name] = current_serial
            current_serial += 1
        course_serial = course_serial_map[course_name]
        for week in weeks:
            if target_week is not None and week != target_week:
                continue
            if week_type == '单' and week % 2 == 0:
                continue
            if week_type == '双' and week % 2 == 1:
                continue
            if week_type_filter == '单' and week % 2 == 0:
                continue
            if week_type_filter == '双' and week % 2 == 1:
                continue
            weekday_map = {"一": 0, "二": 1, "三": 2, "四": 3, "五": 4, "六": 5, "日": 6}
            weekday_num = weekday_map.get(weekday, 0)
            course_date = semester_start + timedelta(weeks=week-1, days=weekday_num)
            date_str = course_date.strftime("%Y-%m-%d")
            date_short = course_date.strftime("%Y-%m-%d")
            for section_num in section_numbers:
                if section_num in section_details:
                    section_name, start_time, end_time = section_details[section_num]
                else:
                    continue
                daily_key = f"{date_str}-{course_name}"
                if daily_key not in daily_course_count:
                    daily_course_count[daily_key] = 1
                else:
                    daily_course_count[daily_key] += 1
                class_count = daily_course_count[daily_key]
                record_id = f"{course_name}-{date_short}-{class_count}"
                classroom_number = current_classroom.replace("产业园C栋", "")
                target_record = {
                    "编号": record_id,
                    "教师账号": teacher_code,
                    "课程及编号": f"{course_name}，{course_serial}",
                    "校区": current_campus,
                    "教学楼": current_building,
                    "教室": classroom_number,
                    "节次类型": season_type,
                    "节次名称": section_name,
                    "开始时间": f"{date_str} {start_time}" if start_time else "",
                    "结束时间": f"{date_str} {end_time}" if end_time else "",
                    "组织机构": "数字技术应用产业学院",
                    "班级": class_name,
                    "简介": "",
                    "是否需要录制": "TRUE",
                    "录制视频是否公开": "false",
                    "是否直播": "",
                    "直播是否公开": ""
                }
                target_data.append(target_record)
    df_target = pd.DataFrame(target_data)
    if not df_target.empty:
        df_target['教室排序'] = df_target['教室'].astype(int)
        df_target = df_target.sort_values(by=['教室排序', '开始时间']).reset_index(drop=True)
        df_target = df_target.drop(columns=['教室排序'])
    df_target.to_excel(output_file, index=False, engine='openpyxl')
    
    # 设置Arial字体和居中对齐
    wb = load_workbook(output_file)
    ws = wb.active
    arial_font = Font(name='Arial')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = arial_font
    wb.save(output_file)
    return len(target_data)

# ==================== 教室班牌文件功能 ====================

# 教室列表（701-706, 708-713, 801-806, 808-813）
CLASSROOMS = []
for floor in ['7', '8']:
    for room in range(1, 14):
        classroom = f"{floor}{room:02d}"
        # 跳过707和807
        if classroom not in ['707', '807']:
            CLASSROOMS.append(classroom)

# 设备ID映射
DEVICE_IDS = {
    '701': 'dc-62-94-0d-d9-88',
    '702': 'dc-62-94-0d-d9-d7',
    '703': 'dc-62-94-0a-8d-db',
    '704': 'dc-62-94-0d-da-93',
    '705': 'dc-62-94-0d-d9-7c',
    '706': 'dc-62-94-0d-d8-44',
    '708': 'dc-62-94-0d-d8-a8',
    '709': 'dc-62-94-0d-d7-32',
    '710': 'dc-62-94-0d-d8-55',
    '711': 'dc-62-94-0d-dc-56',
    '712': 'dc-62-94-0d-d9-13',
    '713': 'dc-62-94-0d-d9-d8',
    '801': 'dc-62-94-0d-d9-96',
    '802': 'dc-62-94-0d-d9-cf',
    '803': 'dc-62-94-0d-d8-2c',
    '804': 'dc-62-94-0d-d7-30',
    '805': 'dc-62-94-0d-da-aa',
    '806': 'dc-62-94-0d-d9-81',
    '808': 'dc-62-94-0d-d9-a2',
    '809': 'dc-62-94-0d-d9-7d',
    '810': 'dc-62-94-0d-db-47',
    '811': 'dc-62-94-0d-da-83',
    '812': 'dc-62-94-0d-d9-91',
    '813': 'dc-62-94-0d-d9-7f'
}

# 星期列表
WEEKDAYS = ['一', '二', '三', '四', '五']

def parse_teacher_info_classroom(teacher_str):
    if pd.isna(teacher_str) or teacher_str == '' or teacher_str == 'nan':
        return ""
    teacher_str = str(teacher_str).strip()
    # 解析方括号中的教师账号和姓名
    match = re.search(r'\[(\d+)\](.+)', teacher_str)
    if match:
        return match.group(2).strip()
    # 保持原有的解析逻辑作为后备
    parts = teacher_str.split('-')
    if len(parts) >= 2:
        return '-'.join(parts[1:]).strip()
    return teacher_str

def parse_section_info_classroom(section_str):
    if pd.isna(section_str) or section_str == '' or section_str == 'nan':
        return {"星期": "", "节次": "", "单双周": ""}
    section_str = str(section_str).strip()
    match = re.match(r'([一二三四五六日])\[(.+?)\](.*)', section_str)
    if match:
        section_content = match.group(2)
        suffix = match.group(3).strip()
        return {
            "星期": match.group(1),
            "节次": section_content,
            "单双周": suffix
        }
    return {"星期": "", "节次": "", "单双周": ""}

def parse_section_range_classroom(section_str):
    sections = []
    if pd.isna(section_str) or section_str == '' or section_str == 'nan':
        return sections
    
    section_str = str(section_str).strip()
    # 移除"节"字
    section_str = section_str.replace('节', '')
    
    parts = section_str.split(',')
    for part in parts:
        part = part.strip()
        if '-' in part:
            try:
                start, end = part.split('-')
                start = int(start)
                end = int(end)
                sections.extend(range(start, end + 1))
            except ValueError:
                pass
        else:
            try:
                sections.append(int(part))
            except ValueError:
                pass
    
    return sections

def read_source_file_classroom(file_path, target_week, season='winter'):
    # 读取源文件
    df = pd.read_excel(file_path, header=None)
    
    # 构建课程数据字典：{(教室, 星期, 节次): {课程信息}}
    course_data = {}
    
    # 当前教室
    current_classroom = None
    
    section_times = get_section_details(season)
    
    for index, row in df.iterrows():
        # 检查是否是教室信息行
        if '产业园C栋' in str(row.iloc[3]) and '容量' in str(row.iloc[13]):
            # 提取教室编号
            classroom_str = str(row.iloc[13])
            match = re.search(r'产业园C栋(\d+)', classroom_str)
            if match:
                classroom_num = match.group(1)
                # 只处理701-713和801-813的教室
                if (classroom_num.startswith('7') or classroom_num.startswith('8')) and len(classroom_num) == 3:
                    current_classroom = classroom_num
                else:
                    current_classroom = None
            else:
                current_classroom = None
            continue
        
        # 跳过前5行（标题和表头）
        if index < 5:
            continue
        
        # 如果没有当前教室，跳过
        if not current_classroom:
            continue
        
        # 尝试获取基本信息
        # 课程信息在第1列（索引1），格式为"[教师账号]课程名称"
        course_cell = row.iloc[1]
        if pd.isna(course_cell):
            continue
        
        # 解析课程信息和教师账号
        course_str = str(course_cell).strip()
        # 提取方括号中的教师账号
        teacher_id_match = re.search(r'\[(\d+)\]', course_str)
        if teacher_id_match:
            teacher_id = teacher_id_match.group(1)
            course_name = course_str.replace(f"[{teacher_id}]", "").strip()
        else:
            teacher_id = ""
            course_name = course_str
        
        # 移除科目中的所有方括号及其内容
        course_name = re.sub(r'\[.*?\]', '', course_name).strip()
        
        # 教师信息在第11列（索引10）
        teacher_cell = row.iloc[10]
        teacher_name = parse_teacher_info_classroom(teacher_cell)
        
        # 班级信息在第14列（索引13）
        class_cell = row.iloc[13]
        class_name = parse_class_info(class_cell)
        
        # 周次在第15列（索引15）
        week_str = row.iloc[15]
        week_info = parse_week_info(week_str)
        
        # 检查是否是目标周
        if target_week not in week_info:
            continue
        
        # 节次在第16列（索引16）
        section_str = row.iloc[16]
        section_info = parse_section_info_classroom(section_str)
        
        # 检查单双周
        week_type = section_info.get('单双周', '')
        if week_type:
            if week_type == '单' and target_week % 2 == 0:
                continue
            if week_type == '双' and target_week % 2 != 0:
                continue
        
        # 解析节次
        sections = parse_section_range_classroom(section_info.get('节次', ''))
        weekday = section_info.get('星期', '')
        
        # 为每个节次创建一条记录
        for section in sections:
            # 只处理1-9节的课程
            if section < 1 or section > 9:
                continue
            
            key = (current_classroom, weekday, section)
            course_data[key] = {
                '班级名称': class_name,
                '设备id': '',
                '星期': weekday,
                '节次': section,
                '时段': f"{section_times.get(section, ('', '', ''))[1]}-{section_times.get(section, ('', '', ''))[2]}",
                '教师': teacher_name,
                '科目': course_name
            }
    
    return course_data

def generate_classroom_sign(course_data, output_path):
    # 生成所有教室、星期、节次的组合
    result_data = []
    
    section_times = get_section_details('winter')  # 用于空记录的时段
    
    for classroom in CLASSROOMS:
        for weekday in WEEKDAYS:
            for section in range(1, 10):  # 1-9节
                key = (classroom, weekday, section)
                if key in course_data:
                    # 使用课程数据
                    record = course_data[key].copy()
                    record['班级名称'] = classroom  # 第一列填教室名称
                    # 填充设备ID
                    record['设备id'] = DEVICE_IDS.get(classroom, '')
                    # 空白字段替换为"空闲"
                    if not record['教师']:
                        record['教师'] = '空闲'
                    if not record['科目']:
                        record['科目'] = '空闲'
                    result_data.append(record)
                else:
                    # 如果没有课程，插入空白记录
                    section_name, start_time, end_time = section_times.get(section, ('', '', ''))
                    result_data.append({
                        '班级名称': classroom,  # 第一列填教室名称
                        '设备id': DEVICE_IDS.get(classroom, ''),
                        '星期': weekday,
                        '节次': section,
                        '时段': f"{start_time}-{end_time}",
                        '教师': '空闲',
                        '科目': '空闲'
                    })
    
    # 创建DataFrame
    df = pd.DataFrame(result_data)
    
    # 重命名列
    df.columns = [
        '班级名称',
        '设备id',
        '星期(一、二、三、四、五、六、日,请按顺序排列)',
        '节次(1、2、3、4、5、6、7、8,请按顺序排列,如果当天中间无课程,请插入一条科目、教师为空白的记录)',
        '时段(09:00-09:45)',
        '教师',
        '科目'
    ]
    
    # 保存Excel文件（不带标题行，后面手动添加）
    df.to_excel(output_path, index=False, sheet_name='Timetable', header=False)
    
    # 加载工作簿进行格式化
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 插入第一行作为标题行
    ws.insert_rows(1)
    
    # 设置标题"教室名称"到"科目"（第2行）
    headers = [
        '班级名称',
        '设备id',
        '星期(一、二、三、四、五、六、日,请按顺序排列)',
        '节次(1、2、3、4、5、6、7、8,请按顺序排列,如果当天中间无课程,请插入一条科目、教师为空白的记录)',
        '时段(09:00-09:45)',
        '教师',
        '科目'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
    
    # 设置Arial字体、加粗、16号字用于标题行
    title_font = Font(name='Arial', bold=True, size=16)
    normal_font = Font(name='Arial')
    
    # 设置第一行的"Timetable"标题（跨列居中）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(row=1, column=1, value='Timetable')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = title_font
    
    # 设置所有单元格的字体和对齐
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = normal_font
    
    # 保存格式化后的文件
    wb.save(output_path)
    
    return len(result_data)

# ==================== 路由 ====================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/config', methods=['GET', 'POST'])
def config_api():
    if request.method == 'GET':
        # 获取配置
        config = load_config()
        return jsonify(config)
    elif request.method == 'POST':
        # 保存配置
        data = request.get_json()
        if data:
            config = load_config()
            config.update(data)
            save_config(config)
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '无效数据'})

@app.route('/generate', methods=['POST'])
def generate():
    if 'file' not in request.files:
        return jsonify({'error': '请选择源文件'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '请选择源文件'}), 400
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'error': '请上传Excel文件 (.xlsx 或 .xls)'}), 400
    try:
        target_week = int(request.form.get('week', 1))
        if target_week < 1:
            return jsonify({'error': '周数必须大于0'}), 400
        
        # 获取转换类型
        convert_type = request.form.get('convert_type', 'classroom_schedule')
        
        # 获取季节类型
        season = request.form.get('season', 'winter')
        if season not in ['winter', 'summer']:
            season = 'winter'
        
        # 获取学期开始日期
        semester_start_date = request.form.get('semester_start_date', '2026-03-09')
        # 验证日期格式
        try:
            datetime.strptime(semester_start_date, "%Y-%m-%d")
        except ValueError:
            return jsonify({'error': '学期开始日期格式不正确，请使用YYYY-MM-DD格式'}), 400
        
        # 保存用户选项到配置
        config = {
            'semester_start_date': semester_start_date,
            'convert_type': convert_type,
            'season': season,
            'week': target_week
        }
        save_config(config)
                
    except ValueError:
        return jsonify({'error': '请输入有效的周数'}), 400
    
    # 保存上传的文件
    temp_input = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    file.save(temp_input.name)
    temp_input.close()
    
    # 生成输出文件
    season_suffix = "_夏季" if season == 'summer' else "_冬季"
    output_filename = f"第{target_week}周课表{season_suffix}.xlsx"
    if convert_type == 'classroom_sign':
        output_filename = f"第{target_week}周班牌{season_suffix}.xlsx"
    temp_output = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    
    try:
        if convert_type == 'classroom_schedule':
            # 生成教室课表
            record_count = convert_classroom_schedule(temp_input.name, temp_output, 
                                                     semester_start_date=semester_start_date,
                                                     target_week=target_week, 
                                                     season=season)
        else:
            # 生成教室班牌
            course_data = read_source_file_classroom(temp_input.name, target_week, season=season)
            record_count = generate_classroom_sign(course_data, temp_output)
        
        # 删除临时输入文件
        os.unlink(temp_input.name)
        
        # 返回下载链接
        return jsonify({
            'success': True,
            'message': f'转换完成！共生成 {record_count} 条记录',
            'filename': output_filename,
            'url': f'/download/{output_filename}'
        })
    except Exception as e:
        # 清理文件
        if os.path.exists(temp_input.name):
            os.unlink(temp_input.name)
        if os.path.exists(temp_output):
            os.unlink(temp_output)
        return jsonify({'error': '转换过程中出现错误: ' + str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return jsonify({'error': '文件不存在'}), 404

if __name__ == '__main__':
    print("课表生成器启动中...")
    print("请在浏览器中访问: http://127.0.0.1:8088")
    print("按 CTRL+C 停止程序")
    app.run(host='127.0.0.1', port=8088, debug=True)